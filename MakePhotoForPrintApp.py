from flask import Flask, render_template, request, redirect, session, jsonify, send_file, Blueprint
from requests_oauthlib import OAuth1Session
import requests
import json
import os
import traceback
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from io import BytesIO
import time
import zipfile


MakePhotoForPrintApp = Flask(__name__)
FMakePhotoForPrintBlueprint = Blueprint('MakePhotoForPrint', __name__)

MakePhotoForPrintApp.config.from_object('Config.Config')


@FMakePhotoForPrintBlueprint.route('/', methods=['GET', 'POST'])
def MakePhotoForPrint():
    global UsersInfo
    if request.method == "POST":
        UsersInfo[len(UsersInfo.keys())+1]={}
        session["UserId"]=max(UsersInfo.keys())

        UsersInfo[session["UserId"]]['OrdersDate']=".".join(reversed(request.form["OrdersDate"].split("-")))

        UsersInfo[session["UserId"]]['OrdersFile'] = request.files['OrdersFile'].read().decode('utf-8').replace('\ufeff', '').replace('\xa0', ' ').strip()

        return redirect('/create_canvas')
    else:
        BasePath = os.getcwd()
        if "UserId" in session and os.path.exists(f'{BasePath}/ResultFiles{session["UserId"]}.zip'):
            os.remove(f'{BasePath}/ResultFiles{session["UserId"]}.zip')
            
        return render_template('MakePhotoForPrint.html')



@FMakePhotoForPrintBlueprint.route('/download_canvas', methods=['GET', 'POST'])
def DownloadCanvas():
    if request.method == 'POST':
        ResultFiles = BytesIO(UsersInfo[session["UserId"]]["ResultFiles"].read())
        UsersInfo[session["UserId"]]["ResultFiles"].seek(0)
        ResultFiles.seek(0) 

        return send_file(ResultFiles, as_attachment=True, download_name="ResultFiles.zip")
    else:
        return render_template('DownloadCanvas.html')



@FMakePhotoForPrintBlueprint.route('/load_log/', methods=['GET', 'POST'])
def LoadLog():
    if session["UserId"] in UsersInfo and 'LogFile' in UsersInfo[session["UserId"]]:
        Log=UsersInfo[session["UserId"]]['LogFile'].split('\n')
        if len(Log)>1:
            Procces=int(((len(Log)-1)/len(json.loads(Log[0])))*100)
        else:
            Procces=0
        return jsonify({'Log': UsersInfo[session["UserId"]]['LogFile'], "Procces":str(Procces)})
    else: return jsonify({'Log': "", "Procces":"100"})


@FMakePhotoForPrintBlueprint.route('/create_canvas', methods=['GET', 'POST'])
def CreateCanvas():
    if request.method == "POST":
        return redirect('/download_canvas')
    else:
        UsersInfo[session["UserId"]]['LogFile'] = ""
        return render_template('CreateCanvas.html')


@FMakePhotoForPrintBlueprint.route('/create_canvas_function/', methods=['GET', 'POST'])
def CreateCanvasFunction():
    #Get Data
    OrdersFile = UsersInfo[session["UserId"]]["OrdersFile"].split('\n')
    OrdersInfo=[OrderInfo.split('\t') for OrderInfo in OrdersFile[1:]]
    OrderIds=[]
    for Number, Data, PrintType in OrdersInfo:
        if Data.replace(' ', '')==UsersInfo[session["UserId"]]['OrdersDate'] and 'дора' not in PrintType.replace(' ', '').lower():
            if 'dtf' in PrintType.replace(' ', '').lower() or 'сувенирка' in PrintType.replace(' ', '').lower():
                OrderIds.append(int(Number.replace(" ", "")))


    UsersInfo[session["UserId"]]['LogFile'] += str(OrderIds)
    
    SizesXlsx = load_workbook(filename=BytesIO(requests.get("https://docs.google.com/spreadsheets/d/1mv6epUufD3Q0UTisHDygZ7qRas8w9fHVIVh2dchitdo/export?format=xlsx").content))
    SizesXlsxSheet=SizesXlsx[SizesXlsx.sheetnames[0]]

    UsersInfo[session["UserId"]]['DownloadImages']={}

    UsersInfo[session["UserId"]]['DownloadPreviewImages']={}

    UsersInfo[session["UserId"]]['DownloadImages3D']={}

    UsersInfo[session["UserId"]]['DownloadPreviewImages3D']={}

    SheetsInfo=[]
    
    for SheetLine in range(1, SizesXlsxSheet.max_row+1):
        SheetsInfo.append([str(SizesXlsxSheet['A'+str(SheetLine)].value).replace(' ', ''), SizesXlsxSheet['B'+str(SheetLine)].value, SizesXlsxSheet['C'+str(SheetLine)].value, SizesXlsxSheet['D'+str(SheetLine)].value, SizesXlsxSheet['E'+str(SheetLine)].value, SizesXlsxSheet['F'+str(SheetLine)].value, SizesXlsxSheet['G'+str(SheetLine)].value])

    MagentoInfo = OAuth1Session(MakePhotoForPrintApp.config["MFEST_ACCES_TOKEN"], client_secret = MakePhotoForPrintApp.config["MFEST_CLIENT_SECRET"], resource_owner_key = MakePhotoForPrintApp.config["MFEST_RESOURCE_OWNER_KEY"], resource_owner_secret = MakePhotoForPrintApp.config["MFEST_RESOURCE_OWNER_SECRET"])
    
    NumberDownload=0
    GetInfoText=[]

    for OrderId in OrderIds:
        try:
            GetInfoText.append('Order Number '+str(OrderId)+':')
            GetInfoText.append('')

            UsersInfo[session["UserId"]]['LogFile'] += "\n"+str(OrderId)

            OrderInfo=MagentoInfo.get("https://pre.mfest.com.ua/rest/V1/orders?searchCriteria[filter_groups][2][filters][0][field]=increment_id&searchCriteria[filter_groups][2][filters][0][value]="+str(OrderId)+"&searchCriteria[filter_groups][2][filters][0][condition_type]=eq")

            SessionRequest = requests.Session()
                    
            for Info in OrderInfo.json()['items'][0]['items'][::2]:
                for ProductCount in range(int(Info['qty_ordered'])):
                    Name=Info['name']
                    FullSKU=Info['sku'].replace(' ', '')
                    
                    ElementExsist=0

                    Result = next((SheetLine for SheetLine in SheetsInfo if SheetLine[0] == FullSKU), None)

                    if Result==None:
                        pass
                    elif Result[4]=='no':
                        ElementExsist=1
                    else:
                        ElementExsist=2

                    
                    if ElementExsist==0:
                        GetInfoText.append("Ім'я: "+Name)
                        GetInfoText.append('Повний SKU: '+FullSKU)
                        GetInfoText.append('Статус: Скасовано. Цього SKU немає в списку')
                        GetInfoText.append('')
                    elif ElementExsist==1:
                        GetInfoText.append("Ім'я: "+Name)
                        GetInfoText.append('Повний SKU: '+FullSKU)
                        GetInfoText.append('Статус: Скасування. Елемент не використовується')
                        GetInfoText.append('')
                    else:
                        DownloadImages = {}
                        for Info2 in OrderInfo.json()['items'][0]['items']:
                            if 'extension_attributes' in Info2 and str(Info['item_id']) in str(Info2['extension_attributes']['design_info']['archive']) and Info2['sku'].replace(' ', '')==FullSKU:
                                ImageNumber=1
                                for ImageBaseUrl, ImagePrintUrl in zip([Base['url'] for Base in Info2['extension_attributes']['design_info']['images'] if '/base/' in Base['url']], [Print['url'] for Print in Info2['extension_attributes']['design_info']['images'] if '/print/' in Print['url']]):
                                    with SessionRequest.get(ImagePrintUrl, stream=True) as DownloadImageRequest:
                                        DownloadImages[f'PrintImage{ImageNumber}'] = BytesIO(DownloadImageRequest.content)
                                    ImageNumber += 1

                                    with SessionRequest.get(ImageBaseUrl, stream=True) as DownloadImageRequest:
                                        DownloadImages[f'BaseImage{ImageNumber}'] = BytesIO(DownloadImageRequest.content)
                                    ImageNumber+=1
                                break
                        time.sleep(0.5)  

                        SortedImagesList=sorted([ImageNames.replace('Base', '').replace('Print', '') for ImageNames in DownloadImages.keys()])
                        
                        ProductWidthsHeightsSidesTypePrints=[]
                        for ImageNumber, ImageName in enumerate(SortedImagesList):
                            if ImageNumber%2==0:
                                ElementIsIncorrect=False
                                ImageSize=[]
                                for SheetLine in SheetsInfo:
                                    if SheetLine[0]==FullSKU:
                                        ImageSize.append([SheetLine[4], SheetLine[5], SheetLine[2], SheetLine[3], SheetLine[1], SheetLine[6]])


                                Picture=Image.open(DownloadImages['Print'+ImageName])
                                PictureWidth, PictureHeight=Picture.size
                                NeededWidth=0
                                for PrintWidth, PrintHeight, Width, Height, Side, TypePrint in ImageSize:
                                    if int(Width)==PictureWidth and int(Height)==PictureHeight:
                                        NeededWidth=round(float(str(PrintWidth).replace(',', '.'))*59.05511811)
                                        NeededHeight=round(float(str(PrintHeight).replace(',', '.'))*59.05511811)
                                        PictureSide=Side
                                        PictureTypePrint=TypePrint
                                        break
                                if NeededWidth!=0 and Picture.getbbox()!=None:
                                    if "Чашка" in FullSKU or 'Khameleon' in FullSKU:
                                        if  Picture.getbbox()[2]-Picture.getbbox()[0]<Picture.size[0]:
                                            NeededHeight=round(float(9)*59.05511811)
                                        else:
                                            NeededHeight=round(float(10)*59.05511811)

                                    TopSpace, BottomSpace, LeftSpace, RightSpace=Picture.getbbox()[1], PictureHeight-Picture.getbbox()[3], Picture.getbbox()[0], PictureWidth-Picture.getbbox()[2]
                                    TopSpace, BottomSpace, LeftSpace, RightSpace=TopSpace/PictureHeight, BottomSpace/PictureHeight, LeftSpace/PictureWidth, RightSpace/PictureWidth
                                    Picture=Picture.crop(Picture.getbbox())
                                    PictureWidth, PictureHeight=Picture.size
                                    if PictureWidth/NeededWidth>=PictureHeight/NeededHeight:
                                        NeededWidth2=round(NeededWidth-(NeededWidth*LeftSpace+NeededWidth*RightSpace))
                                        if "Носки" in FullSKU:
                                            NeededWidth2+=118
                                        Picture=Picture.resize((NeededWidth2, round(NeededWidth2/PictureWidth*PictureHeight)), Image.LANCZOS)
                                    else:
                                        NeededHeight=round(NeededHeight-(NeededHeight*TopSpace+NeededHeight*BottomSpace))
                                        NeededWidth=round(NeededHeight/PictureHeight*PictureWidth)
                                        if "Носки" in FullSKU:
                                            NeededWidth+=118
                                        Picture=Picture.resize((NeededWidth, NeededHeight), Image.LANCZOS)

                                    DownloadImage = BytesIO()
                                    Picture.save(DownloadImage, format='PNG')
                                    DownloadImage.seek(0)

                                    FullSKUEdit=FullSKU.replace('/', '\\')
                                    if '3D' not in FullSKU:
                                        UsersInfo[session["UserId"]]['DownloadImages'][f"{OrderId}|{FullSKUEdit}|{PictureSide}|{PictureTypePrint}|{NumberDownload}"] = DownloadImage
                                    else:
                                        UsersInfo[session["UserId"]]['DownloadImages'][f"{OrderId}|{FullSKUEdit}|{PictureSide}|{PictureTypePrint}|{NumberDownload}"] = DownloadImage

                                    ProductWidthsHeightsSidesTypePrints.append([NeededWidth, NeededHeight, PictureSide, PictureTypePrint])
                                
                                    PreviewPicture=Image.open(DownloadImages['Base'+SortedImagesList[ImageNumber+1]])
                                    PreviewPicture=PreviewPicture.resize((160, int(160/PreviewPicture.size[0]*PreviewPicture.size[1])), Image.LANCZOS)

                                    DownloadPreviewImage = BytesIO()
                                    PreviewPicture.save(DownloadPreviewImage, format='PNG')
                                    DownloadPreviewImage.seek(0)
                                    if '3D' not in FullSKU:
                                        UsersInfo[session["UserId"]]['DownloadPreviewImages'][f"{OrderId}|{FullSKUEdit}|{PictureSide}|{PictureTypePrint}|{NumberDownload}"] = DownloadPreviewImage
                                    else:
                                        UsersInfo[session["UserId"]]['DownloadPreviewImages3D'][f"{OrderId}|{FullSKUEdit}|{PictureSide}|{PictureTypePrint}|{NumberDownload}"] = DownloadPreviewImage


                                    if "Носки" in FullSKU:
                                        NumberDownload+=1

                                        UsersInfo[session["UserId"]]['DownloadImages'][f"{OrderId}|{FullSKUEdit}|{PictureSide}|{PictureTypePrint}|{NumberDownload}"] = DownloadImage
                                        UsersInfo[session["UserId"]]['DownloadPreviewImages'][f"{OrderId}|{FullSKUEdit}|{PictureSide}|{PictureTypePrint}|{NumberDownload}"] = DownloadPreviewImage

                                    NumberDownload+=1
                                elif Picture.getbbox()==None:
                                    GetInfoText.append("Ім'я: "+Name)
                                    GetInfoText.append('Повний SKU: '+FullSKU)
                                    GetInfoText.append('Сторона: '+Side)
                                    GetInfoText.append("Статус: Изображение в товаре пустое!")
                                    GetInfoText.append('')
                                else:
                                    ElementIsIncorrect=True

                        
                        if ElementIsIncorrect==True:
                            GetInfoText.append("Ім'я: "+Name)
                            GetInfoText.append('Повний SKU: '+FullSKU)
                            GetInfoText.append("Статус: В товаре указаны неверные данные")
                            GetInfoText.append('')
                        else:
                            for Width, Height, Side, TypePrint in ProductWidthsHeightsSidesTypePrints:
                                GetInfoText.append("Ім'я: "+Name)
                                GetInfoText.append('Повний SKU: '+FullSKU)
                                GetInfoText.append('Сторона: '+Side)
                                GetInfoText.append('Ширина зображення (см): '+str(round(int(Width)/59.05511811)))
                                GetInfoText.append('Висота зображення (см): '+str(round(int(Height)/59.05511811)))
                                GetInfoText.append('Тип печати изображения: '+str(PictureTypePrint))
                                GetInfoText.append('Статус: Створено успішно')
                                GetInfoText.append('')
                                
        except Exception as Error:
            UsersInfo[session["UserId"]]['LogFile'] += "\n"+str(traceback.format_exc())

            GetInfoText.append("В продукте ошибка на втором этапе: "+str(traceback.format_exc()))
            GetInfoText.append('')


        GetInfoText.append('')
        GetInfoText.append('')

        
                    

    ResultFilesBytes = BytesIO()

    #Third Step To Picture
    with zipfile.ZipFile(ResultFilesBytes, 'a', zipfile.ZIP_DEFLATED) as ResultFiles:
        ResultFiles.writestr("ResultFiles/LOG File.txt", '\n'.join(GetInfoText))

        PictureDTF=Image.new('RGBA', (3729, 17717), color = (255, 255, 255,0))
        PictureDTFDraw=ImageDraw.Draw(PictureDTF)
        PictureSubli=Image.new('RGBA', (3872, 11811), color = (255, 255, 255,0))
        PictureSubliDraw=ImageDraw.Draw(PictureSubli)
        PictureDTFNumber=1
        PictureSubliNumber=1
        EmptySpacesDTF=[]
        EmptySpacesSubli=[]

        while True:
            if all(len(DownloadImageName.split('|')) > 2 for DownloadImageName in UsersInfo[session["UserId"]]['DownloadImages'].keys()):
                break
            time.sleep(0.5)


        def CreatePicture(DownloadImageName, EmptySpacesList, FinishPictureWidth, FinishPicture, FinishPictureFullWidth, FinishPictureDraw, PictureNumber, MaxImageInARow, PictureType, FinishPictureFullHeight):
            OpenPicture=Image.open(UsersInfo[session["UserId"]]['DownloadImages'][DownloadImageName]).convert("RGBA")
            OpenPreviewPicture=Image.open(UsersInfo[session["UserId"]]['DownloadPreviewImages'][DownloadImageName]).convert("RGBA")
            
            if OpenPicture.size[0]>1654:
                OpenPicture = OpenPicture.rotate(-90, expand=True)

            TextInfo=str(DownloadImageName.split('|')[0])+'\n'+str(DownloadImageName.split('|')[1])
            TextSize=str(DownloadImageName.split('|')[-3]).replace('front', 'Перед').replace('back', 'Спина')
            
            PictureText=Image.new('RGBA', (2000, 2000), color = (255, 255, 255,0))
            PictureTextDraw=ImageDraw.Draw(PictureText)
            
            if TextSize=='Спина':
                PictureTextDraw.text((0, 0), TextInfo, fill='black', font=ImageFont.truetype('AmericanTypewriterBold.ttf', size=30), antialias=True)
                PictureTextDraw.text((0, PictureText.getbbox()[3]-PictureText.getbbox()[1] + 7), TextSize, fill='red', font=ImageFont.truetype('AmericanTypewriterBold.ttf', size=70), antialias=True)
            else:
                PictureTextDraw.text((0, 0), TextInfo+f'\n{TextSize}', fill='black', font=ImageFont.truetype('AmericanTypewriterBold.ttf', size=30), antialias=True)
            PictureText=PictureText.crop(PictureText.getbbox())

            if PictureText.size[0]>=OpenPicture.size[0]:
                PictureText=PictureText.resize((OpenPicture.size[0], int((PictureText.size[1]*OpenPicture.size[0])/PictureText.size[0])), Image.LANCZOS)
                
            TextHeight=PictureText.size[1]
            FullPicture=Image.new('RGBA', (OpenPicture.size[0], OpenPicture.size[1]+25+TextHeight), color = (255, 255, 255,0))
            FullPicture.paste(OpenPicture, (0, 0), mask=OpenPicture)
            FullPicture.paste(PictureText, (0, OpenPicture.size[1]+25), mask=PictureText)
            



            UseEmptySpace=None
            for EmptySpaceIndex, EmptySpace in enumerate(reversed(EmptySpacesList)):
                EmptySpaceWidth=FinishPictureWidth-EmptySpace[0][-1][0]
                EmptySpaceHeight=EmptySpace[0][-1][1]
                if FullPicture.size[0]<=EmptySpaceWidth and FullPicture.size[1]<=EmptySpaceHeight:
                    UseEmptySpace=len(EmptySpacesList)-1-EmptySpaceIndex
                    break

            if UseEmptySpace==None:
                FinishPictureHeight=FinishPicture.getbbox()[3] if FinishPicture.getbbox() != None else 0

                if FinishPictureHeight+FullPicture.size[1]>FinishPictureFullHeight:
                    FinishPicture=FinishPicture.crop((0, FinishPicture.getbbox()[1], FinishPictureFullWidth, FinishPicture.getbbox()[3])) 

                    FinishPictureBytes = BytesIO()
                    FinishPicture.save(FinishPictureBytes, format='TIFF', dpi=(150, 150))
                    FinishPictureBytes.seek(0)

                    ResultFiles.writestr(f'ResultFiles/FinishImage{PictureType}{str(PictureNumber)}.tiff', FinishPictureBytes.getvalue())

                    PictureNumber+=1
                    FinishPicture=Image.new('RGBA', (FinishPictureFullWidth, FinishPictureFullHeight), color = (255, 255, 255,0))
                    FinishPictureDraw=ImageDraw.Draw(FinishPicture)
                    EmptySpacesList=[]
                    FinishPictureHeight=FinishPicture.getbbox()[3] if FinishPicture.getbbox() != None else 0
                
                EmptySpacesList.append([ [( int(FullPicture.size[0])+30, int(FullPicture.size[1]) )], FinishPictureHeight])
                FinishPicture.paste(FullPicture, (0, FinishPictureHeight+60), mask=FullPicture)
                FinishPicture.paste(OpenPreviewPicture, (FinishPictureWidth+1, FinishPictureHeight+60), mask=OpenPreviewPicture)
            else:
                EmptySpaceWidth=FinishPictureWidth-EmptySpacesList[UseEmptySpace][0][-1][0]
                EmptySpaceHeight=EmptySpacesList[UseEmptySpace][0][-1][1]
                EmptyHeight=EmptySpacesList[UseEmptySpace][1]

                FinishPicture.paste(FullPicture, (EmptySpacesList[UseEmptySpace][0][-1][0], EmptyHeight+60), mask=FullPicture)
                FinishPicture.paste(OpenPreviewPicture, (FinishPictureWidth+1+(int(len(EmptySpacesList[UseEmptySpace][0]))*160), EmptyHeight+60), mask=OpenPreviewPicture)

                EmptySpacesList[UseEmptySpace][0].append((EmptySpacesList[UseEmptySpace][0][-1][0]+FullPicture.size[0]+30, EmptySpaceHeight))

                if len(EmptySpacesList[UseEmptySpace][0])==MaxImageInARow:
                    EmptySpacesList=EmptySpacesList[:UseEmptySpace]+EmptySpacesList[UseEmptySpace+1:]
                
                

            return [EmptySpacesList, FinishPicture, FinishPictureDraw, PictureNumber]

        for DownloadImageName in sorted(UsersInfo[session["UserId"]]['DownloadImages'].keys(), key=lambda x: x.split('|')[1]):
            if DownloadImageName.split('|')[-2]=="DTF":
                CreatedPicture=CreatePicture(DownloadImageName, EmptySpacesDTF, 3248, PictureDTF, 3729, PictureDTFDraw, PictureDTFNumber, 3, "DTF", 17717)
                EmptySpacesDTF=CreatedPicture[0]
                PictureDTF=CreatedPicture[1]
                PictureDTFDraw=CreatedPicture[2]
                PictureDTFNumber=CreatedPicture[3]
            else:
                CreatedPicture=CreatePicture(DownloadImageName, EmptySpacesSubli, 3071, PictureSubli, 3872, PictureSubliDraw, PictureSubliNumber, 5, "Subli", 11811)
                EmptySpacesSubli=CreatedPicture[0]
                PictureSubli=CreatedPicture[1]
                PictureSubliDraw=CreatedPicture[2]
                PictureSubliNumber=CreatedPicture[3]

            
        if PictureDTF.getbbox()!=None:
            PictureDTF=PictureDTF.crop((0, PictureDTF.getbbox()[1], 3729, PictureDTF.getbbox()[3])) 

            PictureDTFBytes = BytesIO()
            PictureDTF.save(PictureDTFBytes, format='TIFF', dpi=(150, 150))
            PictureDTFBytes.seek(0)

            ResultFiles.writestr(f'ResultFiles/FinishImageDTF{str(PictureDTFNumber)}.tiff', PictureDTFBytes.getvalue())

        if PictureSubli.getbbox()!=None:
            PictureSubli=PictureSubli.crop((0, PictureSubli.getbbox()[1], 3872, PictureSubli.getbbox()[3]))

            PictureSubliBytes = BytesIO()
            PictureSubli.save(PictureSubliBytes, format='TIFF', dpi=(150, 150))
            PictureSubliBytes.seek(0)

            ResultFiles.writestr(f'ResultFiles/FinishImageSubli{str(PictureSubliNumber)}.tiff', PictureSubliBytes.getvalue())

    ResultFilesBytes.seek(0)

    UsersInfo[session["UserId"]]={"ResultFiles":ResultFilesBytes}

    return jsonify({'Log': ""})



if __name__ == '__main__':
    UsersInfo = {}
    MakePhotoForPrintApp.run(debug=False)
    # MakePhotoForPrintApp.run(debug=False, host="127.1.2.140", port=3000)